from selenium import webdriver
import time
import openpyxl
import random

wb = openpyxl.load_workbook('list.xlsx')
ws = wb['Sheet1']

# 금오사이 접속
driver = webdriver.Chrome("./chromedriver")
driver.get("https://sugang.kumoh.ac.kr/html/stud/sugang.html")

#아이디 입력
time.sleep(2)
driver.find_element_by_css_selector("input#Form_로그인\.아이디.mi03").click()
search_box = driver.find_element_by_name('input#Form_로그인.아이디.mi03')
search_box.send_keys("gam_mun@naver.com")

time.sleep(2)

"""
#비밀번호 입력
time.sleep(2)
search_box = driver.find_element_by_name('user_id')
search_box.send_keys("gam_mun@naver.com")
search_box = driver.find_element_by_name('password')
search_box.send_keys("qwqw4313")

#로그인
search_button = driver.find_element_by_css_selector("button.eq.button.button-primary.button-expand.button-rounded.button-shadow.margin-bottom-small.waves-effect")
search_button.click()

pageCount = 1
rowPoint = 2
isEnd = False
writers = []
while not isEnd:
    #강의후기 페이지
    driver.get("https://kumoh42.com/index.php?mid=course_review&page=" + str(pageCount))

    #한페이지의 전체 게시글 수집
    time.sleep(3)
    everyContents = driver.find_element_by_css_selector("div.ed.card.row")
    contents = everyContents.find_elements_by_css_selector("a.ed.overlay.overlay-fill.overlay-top")
    contentURLs = []
    
    for content in contents:
        contentURLs.append(content.get_attribute('href'))
    
    for contentURL in contentURLs:
        #기본정보 긁어오기
        driver.get(contentURL)
        try:
            sortation = driver.find_element_by_css_selector("h5.ed.text-bold.margin-bottom-xsmall").text
        except:
            sortation = "정보없음"
        baseData = driver.find_elements_by_css_selector("td")
        
        writer = driver.find_element_by_css_selector("span.ed.margin-right-xsmall > a.ed.text-bold.link").text
        evaluation = baseData[0].text
        year = baseData[1].text
        semester = baseData[2].text
        name = baseData[3].text
        professor = baseData[4].text
        
        #print
        print(year + "  " + semester + "  " + sortation + "  " + writer + "  " + name + "  " + evaluation + "  " + professor)
        ws.cell(row = rowPoint, column=1).value = year
        ws.cell(row = rowPoint, column=2).value = semester
        ws.cell(row = rowPoint, column=3).value = sortation
        ws.cell(row = rowPoint, column=4).value = writer
        ws.cell(row = rowPoint, column=5).value = name
        ws.cell(row = rowPoint, column=6).value = evaluation
        ws.cell(row = rowPoint, column=7).value = professor
        rowPoint += 1
        
        writers.append(writer)

        if contentURL == "https://kumoh42.com/index.php?mid=course_review&page=5&document_srl=1162315":
            isEnd = True
            break
        
    pageCount += 1
    

print("작성자들")

ws = wb['Sheet2']
rowPoint = 2
for writer in writers:
    print(writer)
    ws.cell(row = rowPoint, column=1).value = writer
    rowPoint += 1

random.shuffle(writers)
selectedWriters = []
rowPoint = 2
for writer in writers:
    if canInsert(writer, selectedWriters):
        selectedWriters.append(writer)
        ws.cell(row = rowPoint, column=7).value = writer
        rowPoint += 1
    
wb.save('reviewWinner.xlsx')
wb.close()
print("종료되었습니다.") 
    
"""